from io import BytesIO

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from starlette.responses import StreamingResponse
from openpyxl.worksheet.table import Table, TableStyleInfo


class ExcelGenerator:
    def generate(self, data):
        df = pd.DataFrame(data)

        df['crawling_date'] = pd.to_datetime(df['crawling_date'])
        df.sort_values(by='crawling_date', ascending=False, inplace=True)

        df.drop_duplicates(subset='url', keep='first', inplace=True)

        df.rename(columns={
            'fabricator': 'Fabricante',
            'model': 'Modelo',
            'year': 'Ano',
            'price': 'Preço',
            'worked_hours': 'Horas',
            'url': 'URL',
            'crawling_date': 'Data da Busca'
        }, inplace=True)

        df['Cód. Modelo'] = ''

        buffer = BytesIO()

        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

        buffer.seek(0)

        book = openpyxl.load_workbook(buffer)
        sheet = book.active

        sheet.freeze_panes = "A2"

        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for i, row in enumerate(sheet.iter_rows(min_row=2),
                                start=2):
            fill = fill_blue if i % 2 == 0 else fill_white
            for cell in row:
                cell.fill = fill

        tab = Table(displayName="Table1", ref=sheet.dimensions)

        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style

        sheet.add_table(tab)

        buffer = BytesIO()
        book.save(buffer)
        buffer.seek(0)

        return StreamingResponse(buffer, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                 headers={'Content-Disposition': 'attachment; filename="dataframe.xlsx"'})